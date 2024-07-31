from openpyxl import Workbook

def create_excel(name: str, data: list, first_column_name: str, second_column_name: str):
    """
    creates a simple excel with 2 columns
    :param name: name of the excel file
    :param data: the list of the info
    :param first_column_name: the name of the first column
    :param second_column_name: the name of the second column
    :return:
    """
    wb = Workbook()
    ws = wb.active
    ws['A1'] = first_column_name
    ws['B1'] = second_column_name
    for row in data:
        ws.append(row)

    wb.save(f'{name}.xlsx')

def create_complex_exccel(name: str, data: list, first_column_name: str= '', second_column_name: str = '', third_column_name: str =''):
    """
    creates an excel with more then 2 columns
    :param name: name of the excel file
    :param data: the list of the info
    :param first_column_name: the name of the first column
    :param second_column_name: the name of the second column
    :param third_column_name: the name of the third column
    :return:
    """
    wb = Workbook()

    ws = wb.active

    ws['A1'] = first_column_name
    ws['B1'] = second_column_name
    ws['C1'] = third_column_name

    for row_idx, row_data in enumerate(data, start=2):  # Rows are 1-indexed in openpyxl
        for col_idx, cell_value in enumerate(row_data, start=1):  # Columns are 1-indexed in openpyxl
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    wb.save(f"{name}.xlsx")
